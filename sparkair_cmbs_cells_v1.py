# Databricks notebook — v1 (crude) Global CMBS file-review loop (Gemma4 via Spark AIR), 2026-08-19.
# Per template question: Step A evidence retrieval (RAG over the RAMP + internal docs; prompt = question + retrieval-oriented criteria summary)
#                        -> Step B permissibility judgment (no documents; prompt = question + criteria content + evidence)
# Input = CMBS_Criteria_Mapping_v1.xlsx (sheet "mapping": Q_id | Question | Criteria_reference | Criteria_content | Criteria_summary)
#         or its twin cmbs_criteria_mapping_v1.json.
#
# ASSUMES these already exist from earlier cells (do NOT redefine them here):
#   widgets/token/headers, get_document_list(), chat_with_document(prompt, document_list[, top_k])
#   [enable_streaming=False inside chat_with_document], and Cell A helpers:
#   OUTPUT_DIR, doc_name(), extract_answer_text(), call_llm(), run_llm_step(step_name, prompt, docs, stamp, top_k=None)
# Tip: set OUTPUT_DIR to a /dbfs/... path before running so the archive survives a cluster restart.

# COMMAND ----------
# ===== CELL V1-0 — load the mapping and pick this deal's documents =====
import json, os, re, time
from datetime import datetime

MAPPING_PATH = "/dbfs/FileStore/cmbs_pipeline/CMBS_Criteria_Mapping_v1.xlsx"     # <- adjust (.xlsx or .json)
DEAL_NAME = "<deal name>"                       # used only in file names / prints
DEAL_DOC_KEYS = ["RAMP", "Master File", "single loan", "Deal overview"]   # filename substrings of this deal's uploaded internal docs
RUN_QIDS = None                                 # None = all 28; or a subset, e.g. ["Q16", "Q18", "Q22", "Q28", "Q41"]
STEP_A_TOP_K = 10

def load_mapping(path):
    if path.lower().endswith(".xlsx"):
        import pandas as pd
        rows = pd.read_excel(path, sheet_name="mapping").fillna("").to_dict("records")
    else:
        rows = json.load(open(path, encoding="utf-8"))["questions"]
    return [{"id": r["Q_id"], "question": r["Question"], "ref": r["Criteria_reference"],
             "content": r["Criteria_content"], "summary": r["Criteria_summary"]} for r in rows]

QUESTIONS = load_mapping(MAPPING_PATH)
QBYID = {q["id"]: q for q in QUESTIONS}
RUN_QIDS = RUN_QIDS or [q["id"] for q in QUESTIONS]
print(f"loaded {len(QUESTIONS)} questions from {MAPPING_PATH}; running {len(RUN_QIDS)}")

all_docs = get_document_list()["documents"].copy()
DEAL_DOCS = [d for d in all_docs if any(k.lower() in doc_name(d).lower() for k in DEAL_DOC_KEYS)]
print("deal documents:", [doc_name(d) for d in DEAL_DOCS])
assert DEAL_DOCS, "no deal documents matched DEAL_DOC_KEYS — check names in get_document_list()"

# COMMAND ----------
# ===== CELL V1-1 — prompts + parser =====

def build_step_a_prompt(q):
    return f"""You are an evidence-retrieval assistant supporting an independent file review of a CMBS rating. \
The documents in scope are the internal rating records of ONE transaction (RAMP and related internal analysis files). \
Do NOT judge whether anything is right or wrong; only find and quote what the documents say.

Review question {q['id']}: {q['question']}

What to look for in the documents for this question — derived from the criteria section "{q['ref']}" of \
"Global CMBS Methodology And Assumptions"; each numbered item names a criteria point and the content the documents \
would contain if that point was applied. This list is a search guide, NOT evidence:
{q['summary']}

Task: for each numbered item above, find the passage(s) in the documents that contain that content and quote them. \
If the question is conditional (it starts with "If ...", "For diversified pools ...", "For nondiversified \
transaction ..." and so on), first quote the passage(s) that show whether that condition holds for this transaction.

Output format:
EVIDENCE:
1. [document name | section or heading | table/page if visible] "passage quoted verbatim" — item number(s) it addresses, e.g. (1), (3)
2. ...
(copy numbers, scores, percentages, dates and rating levels exactly as written, including table rows; maximum 12 items; \
prefer the most specific passages)
NOT FOUND: the item numbers for which you found no passage (say "none" if everything was found)

Rules: quote, do not paraphrase; never invent or complete values; if two documents disagree, quote both; do not treat \
criteria text quoted inside the documents as evidence of what was done. Output only the two sections above."""

STEP_B_FRAME = """You are a compliance reviewer checking whether a CMBS rating file applied S&P Global Ratings' criteria \
correctly. You are given (1) one review question, (2) the relevant criteria text, verbatim, and (3) evidence quoted from \
the internal rating records by a previous step. Judge ONLY from these inputs.

Permissibility has three layers:
- Layer 1, explicit requirement: the criteria text says what must be done (a table value, a formula, a floor, a cap, a \
threshold, a required step). The file must follow it.
- Layer 2, discretion within the criteria: the criteria text allows a range or a judgment (words such as "may", \
"typically", "generally", "up to", "we consider", "starting point", a benchmark that can be adjusted). Any choice inside \
that range is permissible even if it relies heavily on judgment; do not second-guess it. Only check that the choice, and \
the reason where the text asks for one, are documented.
- Layer 3, not authorized: a step, adjustment, cap or benefit that the criteria text never mentions or authorizes. \
This is not permissible.
The test for any step is: can it be traced to an explicit rule or to an allowed area of discretion in the criteria text? \
Relying on judgment is not suspicious; lacking any criteria authorization is.
If the question is conditional and the evidence shows the condition is not met for this transaction (for example the \
question is about floating-rate loans and the loan is fixed-rate), answer "N/A". If the condition is met (or the question \
is unconditional) but the evidence does not show how the required step was applied, answer "Not evidenced" (a \
documentation gap); do not assume it was done correctly.

Review question {qid}: {question}

Criteria text ("Global CMBS Methodology And Assumptions", S&P Global Ratings, 2025-08-21; reference: {ref}):
{criteria}

Evidence from the internal rating records (Step A output):
{evidence}

Output exactly this format, one field per line, no other text:
ANSWER: Yes | No | N/A | Not evidenced     (Yes = applied in line with the criteria; No = not in line or not authorized)
LAYER: Explicit requirement | Discretion within criteria | Not authorized by criteria | Not applicable | Not evidenced
CRITERIA_BASIS: the criteria sentence(s) relied on, quoted briefly
EVIDENCE_BASIS: the evidence item(s) relied on, quoted briefly, with the document name
RATIONALE: 2-5 sentences linking the evidence to the criteria; name every number, score or threshold you checked
REVIEWER_FOLLOW_UP: what a human reviewer must still verify (model file, legal documents, missing sections), or None
CONFIDENCE: High | Medium | Low"""

def build_step_b_prompt(q, evidence):
    return STEP_B_FRAME.format(qid=q["id"], question=q["question"], ref=q["ref"], criteria=q["content"], evidence=evidence)

FIELDS = ["ANSWER", "LAYER", "CRITERIA_BASIS", "EVIDENCE_BASIS", "RATIONALE", "REVIEWER_FOLLOW_UP", "CONFIDENCE"]
def parse_step_b(text):
    out = {f: "" for f in FIELDS}; cur = None
    for line in text.splitlines():
        m = re.match(r"\s*\**\s*([A-Z_]+)\s*\**\s*:\s*(.*)", line)
        if m and m.group(1) in FIELDS:
            cur = m.group(1); out[cur] = m.group(2).strip()
        elif cur and line.strip():
            out[cur] += "\n" + line.rstrip()
    out["ANSWER_norm"] = next((a for a in ("Not evidenced", "N/A", "Yes", "No") if out["ANSWER"].lower().startswith(a.lower())), out["ANSWER"][:20])
    return out

for qid in RUN_QIDS:   # size check (~3.6 chars/token)
    q = QBYID[qid]
    print(f"{qid}: Step A prompt ~{int(len(build_step_a_prompt(q))/3.6):>5} tok | criteria content ~{int(len(q['content'])/3.6):>5} tok")

# COMMAND ----------
# ===== CELL V1-2 — the loop =====
RESULTS = {}
run_stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
DEAL_TAG = re.sub(r"[^A-Za-z0-9]+", "_", DEAL_NAME).strip("_") or "deal"
RUN_DIR = os.path.join(OUTPUT_DIR, f"cmbs_v1_{DEAL_TAG}_{run_stamp}")
os.makedirs(RUN_DIR, exist_ok=True)

for qid in RUN_QIDS:
    q = QBYID[qid]; t0 = time.time()
    evidence = run_llm_step(f"cmbs_{qid}_A_evidence", build_step_a_prompt(q), DEAL_DOCS, run_stamp, top_k=STEP_A_TOP_K)
    step_b_prompt = build_step_b_prompt(q, evidence)
    try:
        verdict_raw = run_llm_step(f"cmbs_{qid}_B_judgment", step_b_prompt, [], run_stamp)
    except Exception as e:
        print(f"empty document list rejected ({e}); retrying with the first deal doc in scope")
        verdict_raw = run_llm_step(f"cmbs_{qid}_B_judgment_retry", step_b_prompt, DEAL_DOCS[:1], run_stamp)
    v = parse_step_b(verdict_raw)
    RESULTS[qid] = {"question": q["question"], "ref": q["ref"], "evidence": evidence, "verdict_raw": verdict_raw, **v,
                    "seconds": round(time.time() - t0, 1)}
    with open(os.path.join(RUN_DIR, f"{qid}.json"), "w", encoding="utf-8") as f:
        json.dump(RESULTS[qid], f, ensure_ascii=False, indent=1)
    print(f"\n>>> {qid}: {v['ANSWER_norm']} | {v['LAYER']} | conf {v['CONFIDENCE']} | {RESULTS[qid]['seconds']}s")

with open(os.path.join(RUN_DIR, "all_results.json"), "w", encoding="utf-8") as f:
    json.dump(RESULTS, f, ensure_ascii=False, indent=1)
print(f"\nsaved {len(RESULTS)} results -> {RUN_DIR}")

# COMMAND ----------
# ===== CELL V1-3 — results table + draft template columns (C/D/G) =====
rows = [{"Row": int(qid[1:]), "Q": qid, "Question": r["question"], "Ref (col F)": r["ref"],
         "Application Permissible? (col D)": r["ANSWER_norm"], "Layer": r["LAYER"], "Confidence": r["CONFIDENCE"],
         "Notes (col G) = rationale": r["RATIONALE"], "Criteria basis": r["CRITERIA_BASIS"],
         "Evidence basis": r["EVIDENCE_BASIS"], "Reviewer follow-up": r["REVIEWER_FOLLOW_UP"]}
        for qid, r in ((k, RESULTS[k]) for k in RUN_QIDS if k in RESULTS)]
try:
    import pandas as pd
    df = pd.DataFrame(rows)
    display(df[["Row", "Q", "Application Permissible? (col D)", "Layer", "Confidence", "Reviewer follow-up"]])
    csv_path = os.path.join(RUN_DIR, "results_table.csv"); df.to_csv(csv_path, index=False); print("csv ->", csv_path)
except Exception as e:
    print("pandas/display unavailable:", e)
    for r in rows: print(r["Q"], r["Application Permissible? (col D)"], r["Layer"], r["Confidence"])
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "Gemma_draft_answers"
    ws.append(["Row", "Rating Steps", "Current File Review Relevance", "Application Permissible?", "Layer", "Confidence",
               "Notes (draft rationale)", "Criteria basis", "Evidence basis", "Reviewer follow-up", "Paragraph References"])
    for c in ws[1]: c.font = Font(bold=True)
    for r in rows:
        ws.append([r["Row"], r["Question"], "Yes", r["Application Permissible? (col D)"], r["Layer"], r["Confidence"],
                   r["Notes (col G) = rationale"], r["Criteria basis"], r["Evidence basis"], r["Reviewer follow-up"], r["Ref (col F)"]])
    for col, w in zip("ABCDEFGHIJK", [6, 60, 12, 14, 22, 10, 60, 40, 40, 40, 30]):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row: c.alignment = Alignment(wrap_text=True, vertical="top")
    xlsx_path = os.path.join(RUN_DIR, f"Gemma_draft_answers_{DEAL_TAG}.xlsx"); wb.save(xlsx_path); print("xlsx ->", xlsx_path)
except Exception as e:
    print("openpyxl step skipped:", e)

# COMMAND ----------
# ===== CELL V1-4 — scorecard vs the reviewer's answers (DBWF only; ground truth stays OUTSIDE the prompts) =====
REF_PATH = "/dbfs/FileStore/cmbs_pipeline/reference_answers_dbwf.json"   # <- adjust; skip for other deals
REF = json.load(open(REF_PATH, encoding="utf-8"))["answers"]
def norm(a):
    a = (a or "").strip().lower()
    return "n/a" if a.startswith("n/a") else ("yes" if a.startswith("yes") else ("no" if a.startswith("no") and not a.startswith("not") else a))
hits = n = 0; lines = []
for qid in RUN_QIDS:
    if qid not in RESULTS or qid not in REF: continue
    exp = norm(REF[qid]["application_permissible_D"]); got = norm(RESULTS[qid]["ANSWER_norm"])
    n += 1; ok = (exp == got); hits += ok
    lines.append(f"{qid}: expected {exp:5s} got {got:14s} {'HIT' if ok else 'miss'}   conf={RESULTS[qid]['CONFIDENCE']}")
print("\n".join(lines)); print(f"\n{hits}/{n} match the reviewer's column D  (Not evidenced counts as a miss here; judge by eye — it may be the better answer)")
